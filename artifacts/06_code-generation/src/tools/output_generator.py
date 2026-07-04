"""申請書生成ツール

AG-002（交通費精算申請エージェント）および AG-003（経費精算申請エージェント）が
HitL承認（OK）後に申請書ドラフト（Excel）を生成・保存するツール関数を提供する。
"""
import logging
import os
from datetime import datetime

import openpyxl
from pydantic import ValidationError
from strands import tool
from strands.types.tools import ToolContext

from handlers.error_handler import ErrorHandler
from models.data_models import (
    GeneralExpenseFormInput,
    TransportExpenseFormInput,
)

_logger = logging.getLogger(__name__)

_TEMPLATE_DIR = "template"
_OUTPUT_DIR = "outputs"


def _load_template(template_filename: str):
    """テンプレート Excel ファイルを openpyxl でロードして返す。

    Returns:
        tuple[bool, workbook | str]: 成功時 (True, wb)、失敗時 (False, エラーメッセージ)
    """
    template_path = os.path.join(_TEMPLATE_DIR, template_filename)
    if not os.path.exists(template_path):
        _logger.error(
            "%s: テンプレートファイル不存在 path=%s",
            template_filename, template_path,
        )
        return (False, "申請書テンプレートが見つかりません。担当部門にお問い合わせください。")
    try:
        wb = openpyxl.load_workbook(template_path)
        return (True, wb)
    except Exception as e:
        _logger.error(
            "%s: テンプレート読み込み失敗 path=%s error=%s",
            template_filename, template_path, e,
        )
        return (False, "申請書テンプレートの読み込みに失敗しました。担当部門にお問い合わせください。")


def _save_file(wb, file_path: str) -> tuple:
    """ワークブックをファイルに保存する。

    Returns:
        tuple[bool, str]: 成功時 (True, "")、失敗時 (False, エラーメッセージ)
    """
    try:
        os.makedirs(os.path.dirname(file_path) or ".", exist_ok=True)
        wb.save(file_path)
        return (True, "")
    except PermissionError as e:
        _logger.error(
            "ファイル書き込み権限エラー error=%s", e, extra={"file_path": file_path}
        )
        return (False, "ファイルへの書き込み権限がありません。")
    except IOError as e:
        _logger.error(
            "ファイルI/Oエラー error=%s", e, extra={"file_path": file_path}
        )
        return (False, "ファイルの書き込みに失敗しました。")
    except Exception as e:
        _logger.error(
            "ファイル保存想定外エラー error=%s", e, extra={"file_path": file_path}
        )
        return (False, "ファイル保存中に予期しないエラーが発生しました。")


@tool(context=True)
def generate_transport_expense_form(
    sections: list,
    total_fare: int,
    tool_context: ToolContext,
) -> dict:
    """交通費精算申請書ドラフト（Excelファイル）を生成して outputs/ ディレクトリに保存する。
    HitL承認（OK）後にのみ呼び出すこと。収集済み情報に含まれないフィールドは補完しない（BRL-04）。

    Args:
        sections (list[dict]): 移動区間リスト（1件以上）。各要素は以下のキーを持つ辞書:
            - travel_date (str): 移動日（YYYY-MM-DD形式）
            - departure (str): 出発地（正規化済み）
            - destination (str): 目的地（正規化済み）
            - transport_type (str): 交通手段（電車/バス/タクシー/飛行機）
            - fare (int): 交通費（円、非負整数）
            - purpose (str): 業務目的（500文字以内）
        total_fare (int): 交通費合計（円、非負整数）。各区間fareの合計。

    Returns:
        dict: 以下のキーを持つ辞書:
            - file_path (str): 生成した申請書ドラフトのファイルパス（outputs/ 配下）
            - status (str): 生成結果ステータス（"success" または "failure"）
            エラー時: {"success": False, "message": "エラーメッセージ文字列"}
    """
    _logger.info("generate_transport_expense_form: 開始 sections=%d", len(sections) if sections else 0)

    applicant_name = tool_context.invocation_state.get("applicant_name", "")
    application_date = tool_context.invocation_state.get("application_date", "")

    try:
        validated = TransportExpenseFormInput(
            applicant_name=applicant_name,
            application_date=application_date,
            sections=sections,
            total_fare=total_fare,
        )
    except ValidationError as e:
        msg = ErrorHandler.handle_validation_error(e)
        return {"success": False, "message": msg}

    ok, result = _load_template("交通費精算申請書テンプレート.xlsx")
    if not ok:
        return {"success": False, "message": result}
    wb = result

    ws = wb.active
    ws["B3"] = validated.applicant_name
    ws["B4"] = validated.application_date.strftime("%Y-%m-%d")

    for i, section in enumerate(validated.sections):
        row = 7 + i
        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = section.travel_date.strftime("%Y-%m-%d")
        ws[f"C{row}"] = section.departure
        ws[f"D{row}"] = section.destination
        ws[f"E{row}"] = section.transport_type
        ws[f"F{row}"] = section.fare
        ws[f"G{row}"] = section.purpose
        ws[f"H{row}"] = ""

    n = len(validated.sections)
    ws[f"H{7+n+2}"] = validated.total_fare

    output_path = os.path.join(
        _OUTPUT_DIR,
        f"交通費精算申請書_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
    )
    ok, err_msg = _save_file(wb, output_path)
    if not ok:
        return {"success": False, "message": err_msg}

    _logger.info("generate_transport_expense_form: 生成成功 file_path=%s", output_path)
    return {"file_path": output_path, "status": "success"}


@tool(context=True)
def generate_general_expense_form(
    items: list,
    total_amount: int,
    tool_context: ToolContext,
) -> dict:
    """経費精算申請書ドラフト（Excelファイル）を生成して outputs/ ディレクトリに保存する。
    HitL承認（OK）後にのみ呼び出すこと。収集済み情報に含まれないフィールドは補完しない（BRL-04）。

    Args:
        items (list[dict]): 経費リスト（1件以上）。各要素は以下のキーを持つ辞書:
            - purchase_date (str): 購入日（YYYY-MM-DD形式）
            - store_name (str): 店舗名（500文字以内）
            - item_name (str): 品目（500文字以内）
            - expense_category (str): 経費区分（事務用品費/宿泊費/資格精算費/その他経費）
            - amount (int): 金額（円、非負整数）
            - purpose (str): 業務目的（500文字以内）
        total_amount (int): 経費合計（円、非負整数）。各経費amountの合計。

    Returns:
        dict: 以下のキーを持つ辞書:
            - file_path (str): 生成した申請書ドラフトのファイルパス（outputs/ 配下）
            - status (str): 生成結果ステータス（"success" または "failure"）
            エラー時: {"success": False, "message": "エラーメッセージ文字列"}
    """
    _logger.info("generate_general_expense_form: 開始 items=%d", len(items) if items else 0)

    applicant_name = tool_context.invocation_state.get("applicant_name", "")
    application_date = tool_context.invocation_state.get("application_date", "")

    try:
        validated = GeneralExpenseFormInput(
            applicant_name=applicant_name,
            application_date=application_date,
            items=items,
            total_amount=total_amount,
        )
    except ValidationError as e:
        msg = ErrorHandler.handle_validation_error(e)
        return {"success": False, "message": msg}

    ok, result = _load_template("経費精算申請書テンプレート.xlsx")
    if not ok:
        return {"success": False, "message": result}
    wb = result

    ws = wb.active
    ws["B3"] = validated.applicant_name
    ws["B4"] = validated.application_date.strftime("%Y-%m-%d")

    for i, item in enumerate(validated.items):
        row = 7 + i
        ws[f"A{row}"] = i + 1
        ws[f"B{row}"] = item.purchase_date.strftime("%Y-%m-%d")
        ws[f"C{row}"] = item.store_name
        ws[f"D{row}"] = item.item_name
        ws[f"E{row}"] = item.expense_category
        ws[f"F{row}"] = item.amount
        ws[f"G{row}"] = item.purpose
        ws[f"H{row}"] = ""

    n = len(validated.items)
    ws[f"H{7+n+2}"] = validated.total_amount

    output_path = os.path.join(
        _OUTPUT_DIR,
        f"経費精算申請書_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx",
    )
    ok, err_msg = _save_file(wb, output_path)
    if not ok:
        return {"success": False, "message": err_msg}

    _logger.info("generate_general_expense_form: 生成成功 file_path=%s", output_path)
    return {"file_path": output_path, "status": "success"}
